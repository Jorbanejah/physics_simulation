import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from scipy.optimize import fsolve
from openpyxl import load_workbook, Workbook

wb =load_workbook('projectile_motion_data.xlsx')
ws = wb.active

def read_column_and_data_values(ws, nombre_columna, filas):
    col = None
    for cell in ws[1]:
        if cell.value == nombre_columna:
            col = cell.column_letter
            break

    if col is None:
        raise ValueError(f"Does not exist column '{nombre_columna}'")

    return [ws[f"{col}{i}"].value for i in range(2, filas + 1)]


# k= 0
x_0  = read_column_and_data_values(ws, "x_a_k=0", 433)
y_0  = read_column_and_data_values(ws, "y_a_k=0", 433)
vx_0 = read_column_and_data_values(ws, "vx_a_k=0", 433)
vy_0 = read_column_and_data_values(ws, "vy_a_k=0", 433)

x_num_0 = read_column_and_data_values(ws, "x_n_k=0", 433)
y_num_0 = read_column_and_data_values(ws, "y_n_k=0", 433)
vx_num_0 = read_column_and_data_values(ws, "vx_n_k=0", 433)
vy_num_0 = read_column_and_data_values(ws, "vy_n_k=0", 433)

# k= 0.1
x_01 = read_column_and_data_values(ws, "x_a_k=0.1", 400)
y_01 = read_column_and_data_values(ws, "y_a_k=0.1", 400)
vx_01 = read_column_and_data_values(ws, "vx_a_k=0.1", 400)
vy_01 = read_column_and_data_values(ws, "vy_a_k=0.1", 400)

x_num_01 = read_column_and_data_values(ws, "x_n_k=0.1", 400)
y_num_01 = read_column_and_data_values(ws, "y_n_k=0.1", 400)
vx_num_01 = read_column_and_data_values(ws, "vx_n_k=0.1", 400)
vy_num_01 = read_column_and_data_values(ws, "vy_n_k=0.1", 400)

t = read_column_and_data_values(ws, "t", 433)
t1 = read_column_and_data_values(ws, "t", 400)
m = 0.8

def energy(m, g, x, y, vx, vy, E0=None):
    """
    - E_mec: total mechanic energy
    - E_k: kinetic energy
    - E_p: potencial energy
    - P_ext: losses
    """
    x = np.array([float(v) if v not in (None, '') else 0 for v in x])
    y = np.array([float(v) if v not in (None, '') else 0 for v in y])
    vx = np.array([float(v) if v not in (None, '') else 0 for v in vx])
    vy = np.array([float(v) if v not in (None, '') else 0 for v in vy])

    Ek = 0.5 * m * (vx**2 + vy**2)
    Ep = m * g * y 
    Em = Ek + Ep

    if E0 is None:
        E0 = Em[0] if hasattr(Em, '__len__') else Em

    P_ext = E0 - Em
    return Em, Ek, Ep, P_ext


######
# Plot energies over time
######
g = 9.81
Em_0, Ek_0, Ep_0, P_ext_0 = energy(m, g, x_0, y_0, vx_0, vy_0)
Em_0_num, Ek_0_num, Ep_0_num, P_ext_0_num = energy(m, g, x_num_0, y_num_0, vx_num_0, vy_num_0, E0=Em_0[0])
Em_01, Ek_01, Ep_01, P_ext_01 = energy(m, g, x_01, y_01, vx_01, vy_01, E0=Em_0[0])  
Em_01_num, Ek_01_num, Ep_01_num, P_ext_01_num = energy(m, g, x_num_01, y_num_01, vx_num_01, vy_num_01, E0=Em_0[0])

fig, axs = plt.subplots(2, 2, figsize=(12, 8))
axs[0, 0].plot(t, Em_0, label='E_mec k=0')
axs[0, 0].plot(t, Em_0_num, label='E_mec num k=0')
axs[0, 0].set_title('Mechanic energy vs time for k=0')

axs[0, 1].plot(t1, Em_01, label='E_mec k=0.1')
axs[0, 1].plot(t1, Em_01_num, label='E_mec num k=0.1')
axs[0, 1].set_title('Mechanic energy vs time for k=0.1')

axs[1, 0].plot(t, Ek_0, label='E_k')
axs[1, 0].plot(t, Ek_0_num, label='E_k num')
axs[1, 0].plot(t, Ep_0, label='E_p')
axs[1, 0].plot(t, Ep_0_num, label='E_p num')
axs[1, 0].set_title('Kinetic and potential energy vs time for k=0')

axs[1, 1].plot(t1, Ek_01, label='E_k ')
axs[1, 1].plot(t1, Ep_01, label='E_p ')
axs[1, 1].plot(t1, Ek_01_num, label='E_k num')
axs[1, 1].plot(t1, Ep_01_num, label='E_p num')
axs[1, 1].set_title('Kinetic and potential energy vs time for k=0.1')

plt.savefig('C:\\Users\\JORGE\\Desktop\\Programas\\Python\\physics_simulation\\Movimiento_de_proyectiles\\energy.png')



#                                                 Analitical method - Range, Height and time
# Everybody knows how to get the range and height of a parabolic motion. Through the parabolic equation when y = 0 we can solve it and see that 
# R = x = self.v0 **2 * np.sin(2*alpha) / self.g and then we can get the time. On the other hand, when velocity vy = 0 we can solve the equation put inside y equation and take it, too: H = self.v0 **2 * np.sin(alpha)**2 / (2*self.g).
# However, now we are trying to figure out how to deal with parabolic equation through different medium. So, putting y = 0 we can see a trascental equation that we can solve by numerical method (for example 
# plotting both parts right and left, and show where interline) or doing some approximation such as exponential Taylor expansion.

g = 9.81 #Gravity
x0, y0, grades, v0 = 0, 0, 45, 30 # Initial condition
alpha = np.radians(grades)
v0x = v0*np.cos(alpha)
v0y = v0*np.sin(alpha)

T_aprox = [] # Approx time
R_aprox = [] # Approx range


T = []
R = []
H = []
k1 = list(np.linspace(0, 0.8, 20))

for k in k1:
    
    if k == 0:
        T_height = v0y/g
        T_aprox.append(2*v0y/g)
        R_aprox.append(v0x*(2*v0y/g))
        
        H.append(y0 + v0y*T_height - 1/2*g*T_height**2)
        T.append(2*v0y/g)
        R.append(v0x*(2*v0y/g))

    else:
        T_height = -1/k*np.log(g/(k*v0y+g))

        T_aprox.append(2*v0y/g*(1 - k*v0y / (3*g)))
        R_aprox.append(v0x*(T_aprox[-1] - 0.5*k*T_aprox[-1]**2))

        #Solving transcendental equation for T -- numerically
        func = lambda t: t - (k*v0y+g)/(g*k)*(1 - np.exp(-k*t))
        T_real = fsolve(func, T_aprox[-1])[0]
        T.append(T_real)
        R.append(v0x/k*(1- np.exp(-k*T_real)))
        H.append((k*v0y + g)/k**2 * (1 - np.exp(-k*T_height)) - g*T_height/k)



plt.figure(1)
plt.plot(k1, T, label='Real time')
plt.plot(k1, T_aprox, '--',label='Approximated time')
plt.title('Time in air vs resistance')
plt.xlabel('k (1/s)')
plt.ylabel('Time (s)')
plt.legend()
plt.savefig('C:\\Users\\JORGE\\Desktop\\Programas\\Python\\physics_simulation\\Movimiento_de_proyectiles\\time.png')


plt.figure(2)
plt.plot(k1, R, label='Real range')
plt.plot(k1, R_aprox, '--',label='Approximated range')
plt.title('Maximum range vs resistance')
plt.xlabel('k (1/s)')
plt.ylabel('Range (m)')
plt.legend()
plt.savefig('C:\\Users\\JORGE\\Desktop\\Programas\\Python\\physics_simulation\\Movimiento_de_proyectiles\\range.png')


plt.figure(3)
plt.plot(k1, H, '*',label='Height')
plt.xlabel('k (1/s)')
plt.ylabel('Height (m)')
plt.title('Maximum height vs resistance')
plt.legend()
plt.savefig('C:\\Users\\JORGE\\Desktop\\Programas\\Python\\physics_simulation\\Movimiento_de_proyectiles\\height.png')